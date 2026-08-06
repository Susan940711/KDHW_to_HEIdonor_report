import argparse
from collections import defaultdict
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


def to_number(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return 0.0
        text = text.replace(",", "")
        try:
            return float(text)
        except ValueError:
            return 0.0
    return 0.0


def get_header_index(headers: List[str], names: List[str]) -> dict:
    normalized = {name.strip().lower(): idx for idx, name in enumerate(headers)}
    result = {}
    for name in names:
        result[name] = normalized.get(name.strip().lower())
    return result


def build_township_breakdown(source_ws, target_ws) -> None:
    headers = [cell.value if cell.value is not None else "" for cell in next(source_ws.iter_rows(min_row=1, max_row=1))]
    header_map = get_header_index(headers, [
        "Year",
        "period",
        "Project Name",
        "Twp_MIMU",
        "Penta1_U1",
        "Penta1_U5",
        "Penta3_U1",
        "Penta3_U5",
        "MMR1_U1",
        "MMR1_U5",
        "MMR2_U1",
        "MMR2_U5",
        "CD_U1",
        "CD_U5",
        "Td2",
    ])

    required_headers = [
        "Year",
        "period",
        "Project Name",
        "Twp_MIMU",
        "Penta1_U1",
        "Penta1_U5",
        "Penta3_U1",
        "Penta3_U5",
        "MMR1_U1",
        "MMR1_U5",
        "MMR2_U1",
        "MMR2_U5",
        "CD_U1",
        "CD_U5",
        "Td2",
    ]
    for name in required_headers:
        if header_map.get(name) is None:
            raise ValueError(f"Required column '{name}' was not found in the source disaggregate sheet")

    rows_by_key = defaultdict(list)
    for row in source_ws.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue

        year = row[header_map["Year"]] if header_map["Year"] is not None and header_map["Year"] < len(row) else None
        period = row[header_map["period"]] if header_map["period"] is not None and header_map["period"] < len(row) else None
        project_name = row[header_map["Project Name"]] if header_map["Project Name"] is not None and header_map["Project Name"] < len(row) else ""
        twp_mimu = row[header_map["Twp_MIMU"]] if header_map["Twp_MIMU"] is not None and header_map["Twp_MIMU"] < len(row) else ""
        if year is None or period is None or not project_name or not twp_mimu:
            continue

        rows_by_key[(str(year), str(project_name), str(twp_mimu))].append((str(period), row))

    header_row = [
        "Year",
        "Period",
        "Project Name",
        "Twp_MIMU",
        "Penta 1",
        "Penta3",
        "MMR1",
        "MMR2",
        "CD",
        "Td2",
    ]
    target_ws.append(header_row)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in target_ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for (year, project_name, twp_mimu), rows in sorted(rows_by_key.items()):
        quarter_rows = {
            "Q1": [entry for entry in rows if str(entry[0]).upper().startswith("Q1")],
            "Q2": [entry for entry in rows if str(entry[0]).upper().startswith("Q2")],
            "Q3": [entry for entry in rows if str(entry[0]).upper().startswith("Q3")],
            "Q4": [entry for entry in rows if str(entry[0]).upper().startswith("Q4")],
        }

        def sum_metric(metric_suffix: str, rows_for_period: List[tuple]) -> float:
            total = 0.0
            for _, row in rows_for_period:
                col1 = header_map[metric_suffix + "_U1"]
                col2 = header_map[metric_suffix + "_U5"]
                if col1 is not None and col1 < len(row):
                    total += to_number(row[col1])
                if col2 is not None and col2 < len(row):
                    total += to_number(row[col2])
            return total

        def append_period(period_name: str, selected_rows: List[tuple]) -> None:
            penta1 = sum_metric("Penta1", selected_rows)
            penta3 = sum_metric("Penta3", selected_rows)
            mmr1 = sum_metric("MMR1", selected_rows)
            mmr2 = sum_metric("MMR2", selected_rows)
            cd = sum_metric("CD", selected_rows)
            td2 = 0.0
            for _, row in selected_rows:
                col = header_map["Td2"]
                if col is not None and col < len(row):
                    td2 += to_number(row[col])
            target_ws.append([year, period_name, project_name, twp_mimu, penta1, penta3, mmr1, mmr2, cd, td2])

        append_period(f"S1_{year}", quarter_rows["Q1"] + quarter_rows["Q2"])
        append_period(f"S2_{year}", quarter_rows["Q3"] + quarter_rows["Q4"])
        append_period(f"Annual_{year}", rows)


def build_semester_report(input_path: Path, output_path: Path, source_sheet: str = "Indicator", target_sheet: str = "Semester_Report") -> Path:
    wb = load_workbook(input_path, data_only=False)
    if source_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{source_sheet}' not found in {input_path}")

    source_ws = wb[source_sheet]
    headers = [cell.value if cell.value is not None else "" for cell in next(source_ws.iter_rows(min_row=1, max_row=1))]
    header_map = get_header_index(headers, [
        "Year",
        "Organization",
        "Organization ",
        "Project Name",
        "indicator",
        "Q1 Target",
        "Q1 U1 Male",
        "Q1 U1 Female",
        "Q1 1-5 Male ",
        "Q1 1-5 Female",
        "Q1 Total",
        "Q2 Target",
        "Q2 U1 Male",
        "Q2 U1 Female",
        "Q2 1-5 Male ",
        "Q2 1-5 Female",
        "Q2 Total",
        "Q3 Target",
        "Q3 U1 Male",
        "Q3 U1 Female",
        "Q3 1-5 Male ",
        "Q3 1-5 Female",
        "Q3 Total",
        "Q4 Target",
        "Q4 U1 Male",
        "Q4 U1 Female",
        "Q4 1-5 Male ",
        "Q4 1-5 Female",
        "Q4 Total",
    ])

    required_headers = [
        "Year",
        "Organization",
        "Project Name",
        "indicator",
        "Q1 Target",
        "Q1 U1 Male",
        "Q1 U1 Female",
        "Q1 1-5 Male ",
        "Q1 1-5 Female",
        "Q1 Total",
        "Q2 Target",
        "Q2 U1 Male",
        "Q2 U1 Female",
        "Q2 1-5 Male ",
        "Q2 1-5 Female",
        "Q2 Total",
        "Q3 Target",
        "Q3 U1 Male",
        "Q3 U1 Female",
        "Q3 1-5 Male ",
        "Q3 1-5 Female",
        "Q3 Total",
        "Q4 Target",
        "Q4 U1 Male",
        "Q4 U1 Female",
        "Q4 1-5 Male ",
        "Q4 1-5 Female",
        "Q4 Total",
    ]
    for name in required_headers:
        if header_map.get(name) is None:
            raise ValueError(f"Required column '{name}' was not found in sheet '{source_sheet}'")

    if target_sheet in wb.sheetnames:
        del wb[target_sheet]

    target_ws = wb.create_sheet(title=target_sheet)
    header_row = [
        "Year",
        "Organization",
        "Project Name",
        "indicator",
        "S1 Target",
        "S1 Male",
        "S1 Female",
        "S1 Total",
        "S2 Target",
        "S2 Male",
        "S2 Female",
        "S2 Total",
        "Annual Target",
        "Annual Male",
        "Annual Female",
        "Annual Total",
    ]
    target_ws.append(header_row)

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in target_ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in source_ws.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue

        year = row[header_map["Year"]] if header_map["Year"] is not None and header_map["Year"] < len(row) else ""
        organization = row[header_map["Organization"]] if header_map["Organization"] is not None and header_map["Organization"] < len(row) else None
        if organization is None and header_map.get("Organization ") is not None and header_map["Organization "] < len(row):
            organization = row[header_map["Organization "]]
        project_name = row[header_map["Project Name"]] if header_map["Project Name"] is not None and header_map["Project Name"] < len(row) else ""
        indicator = row[header_map["indicator"]] if header_map["indicator"] is not None and header_map["indicator"] < len(row) else ""

        q1_target = to_number(row[header_map["Q1 Target"]] if header_map["Q1 Target"] is not None and header_map["Q1 Target"] < len(row) else None)
        q1_u1_male = to_number(row[header_map["Q1 U1 Male"]] if header_map["Q1 U1 Male"] is not None and header_map["Q1 U1 Male"] < len(row) else None)
        q1_u1_female = to_number(row[header_map["Q1 U1 Female"]] if header_map["Q1 U1 Female"] is not None and header_map["Q1 U1 Female"] < len(row) else None)
        q1_1_5_male = to_number(row[header_map["Q1 1-5 Male "]] if header_map["Q1 1-5 Male "] is not None and header_map["Q1 1-5 Male "] < len(row) else None)
        q1_1_5_female = to_number(row[header_map["Q1 1-5 Female"]] if header_map["Q1 1-5 Female"] is not None and header_map["Q1 1-5 Female"] < len(row) else None)
        q1_total = to_number(row[header_map["Q1 Total"]] if header_map["Q1 Total"] is not None and header_map["Q1 Total"] < len(row) else None)

        q2_target = to_number(row[header_map["Q2 Target"]] if header_map["Q2 Target"] is not None and header_map["Q2 Target"] < len(row) else None)
        q2_u1_male = to_number(row[header_map["Q2 U1 Male"]] if header_map["Q2 U1 Male"] is not None and header_map["Q2 U1 Male"] < len(row) else None)
        q2_u1_female = to_number(row[header_map["Q2 U1 Female"]] if header_map["Q2 U1 Female"] is not None and header_map["Q2 U1 Female"] < len(row) else None)
        q2_1_5_male = to_number(row[header_map["Q2 1-5 Male "]] if header_map["Q2 1-5 Male "] is not None and header_map["Q2 1-5 Male "] < len(row) else None)
        q2_1_5_female = to_number(row[header_map["Q2 1-5 Female"]] if header_map["Q2 1-5 Female"] is not None and header_map["Q2 1-5 Female"] < len(row) else None)
        q2_total = to_number(row[header_map["Q2 Total"]] if header_map["Q2 Total"] is not None and header_map["Q2 Total"] < len(row) else None)

        q3_target = to_number(row[header_map["Q3 Target"]] if header_map["Q3 Target"] is not None and header_map["Q3 Target"] < len(row) else None)
        q3_u1_male = to_number(row[header_map["Q3 U1 Male"]] if header_map["Q3 U1 Male"] is not None and header_map["Q3 U1 Male"] < len(row) else None)
        q3_u1_female = to_number(row[header_map["Q3 U1 Female"]] if header_map["Q3 U1 Female"] is not None and header_map["Q3 U1 Female"] < len(row) else None)
        q3_1_5_male = to_number(row[header_map["Q3 1-5 Male "]] if header_map["Q3 1-5 Male "] is not None and header_map["Q3 1-5 Male "] < len(row) else None)
        q3_1_5_female = to_number(row[header_map["Q3 1-5 Female"]] if header_map["Q3 1-5 Female"] is not None and header_map["Q3 1-5 Female"] < len(row) else None)
        q3_total = to_number(row[header_map["Q3 Total"]] if header_map["Q3 Total"] is not None and header_map["Q3 Total"] < len(row) else None)

        q4_target = to_number(row[header_map["Q4 Target"]] if header_map["Q4 Target"] is not None and header_map["Q4 Target"] < len(row) else None)
        q4_u1_male = to_number(row[header_map["Q4 U1 Male"]] if header_map["Q4 U1 Male"] is not None and header_map["Q4 U1 Male"] < len(row) else None)
        q4_u1_female = to_number(row[header_map["Q4 U1 Female"]] if header_map["Q4 U1 Female"] is not None and header_map["Q4 U1 Female"] < len(row) else None)
        q4_1_5_male = to_number(row[header_map["Q4 1-5 Male "]] if header_map["Q4 1-5 Male "] is not None and header_map["Q4 1-5 Male "] < len(row) else None)
        q4_1_5_female = to_number(row[header_map["Q4 1-5 Female"]] if header_map["Q4 1-5 Female"] is not None and header_map["Q4 1-5 Female"] < len(row) else None)
        q4_total = to_number(row[header_map["Q4 Total"]] if header_map["Q4 Total"] is not None and header_map["Q4 Total"] < len(row) else None)

        s1_target = q1_target + q2_target
        s1_male = q1_u1_male + q1_1_5_male + q2_u1_male + q2_1_5_male
        s1_female = q1_u1_female + q1_1_5_female + q2_u1_female + q2_1_5_female
        s1_total = q1_total + q2_total
        s2_target = q3_target + q4_target
        s2_male = q3_u1_male + q3_1_5_male + q4_u1_male + q4_1_5_male
        s2_female = q3_u1_female + q3_1_5_female + q4_u1_female + q4_1_5_female
        s2_total = q3_total + q4_total
        annual_target = q1_target + q2_target + q3_target + q4_target
        annual_male = q1_u1_male + q1_1_5_male + q2_u1_male + q2_1_5_male + q3_u1_male + q3_1_5_male + q4_u1_male + q4_1_5_male
        annual_female = q1_u1_female + q1_1_5_female + q2_u1_female + q2_1_5_female + q3_u1_female + q3_1_5_female + q4_u1_female + q4_1_5_female
        annual_total = q1_total + q2_total + q3_total + q4_total

        row_values = [
            year,
            organization,
            project_name,
            indicator,
            s1_target,
            s1_male,
            s1_female,
            s1_total,
            s2_target,
            s2_male,
            s2_female,
            s2_total,
            annual_target,
            annual_male,
            annual_female,
            annual_total,
        ]
        target_ws.append(row_values)
        current_row = target_ws.max_row
        target_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        total_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
        for column_index in [5, 9, 13]:
            cell = target_ws.cell(row=current_row, column=column_index)
            cell.fill = target_fill
        for column_index in [8, 12, 16]:
            cell = target_ws.cell(row=current_row, column=column_index)
            cell.fill = total_fill

    disaggregate_source_sheet = "VTHC_Doses disaggregate"
    if disaggregate_source_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{disaggregate_source_sheet}' not found in {input_path}")

    if "twn_breakdown" in wb.sheetnames:
        del wb["twn_breakdown"]
    breakdown_ws = wb.create_sheet(title="twn_breakdown")
    build_township_breakdown(wb[disaggregate_source_sheet], breakdown_ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Transform the Indicator sheet into a semester report sheet in an Excel workbook")
    parser.add_argument("input_workbook", nargs="?", default=r"C:\Users\mssam\OneDrive\Data automation\KDHW\quarterly compile.update.xlsx")
    parser.add_argument("output_workbook", nargs="?", default=None)
    parser.add_argument("--source-sheet", default="Indicator")
    parser.add_argument("--target-sheet", default="Semester_Report")
    args = parser.parse_args()

    input_path = Path(args.input_workbook)
    if args.output_workbook:
        output_path = Path(args.output_workbook)
    else:
        output_path = input_path.with_name(input_path.stem + "_semester_report.xlsx")

    result = build_semester_report(input_path=input_path, output_path=output_path, source_sheet=args.source_sheet, target_sheet=args.target_sheet)
    print(f"Created {result}")


if __name__ == "__main__":
    main()
