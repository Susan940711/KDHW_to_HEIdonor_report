from pathlib import Path
import tempfile
from openpyxl import Workbook, load_workbook
from transform_indicator_to_semester_report import build_semester_report

with tempfile.TemporaryDirectory() as tmpdir:
    tmpdir_path = Path(tmpdir)
    input_path = tmpdir_path / 'sample.xlsx'
    output_path = tmpdir_path / 'out.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Indicator'
    ws.append(['Year','Organization ','Project Name','indicator','Q1 Target','Q1 U1 Male','Q1 U1 Female','Q1 1-5 Male ','Q1 1-5 Female','Q1 Total','Q2 Target','Q2 U1 Male','Q2 U1 Female','Q2 1-5 Male ','Q2 1-5 Female','Q2 Total','Q3 Target','Q3 U1 Male','Q3 U1 Female','Q3 1-5 Male ','Q3 1-5 Female','Q3 Total','Q4 Target','Q4 U1 Male','Q4 U1 Female','Q4 1-5 Male ','Q4 1-5 Female','Q4 Total'])
    ws.append([2025,'KDHW','REACH-KK','Penta3 under 1-yr-old',400,10,12,3,4,19,400,20,22,5,6,33,400,30,28,7,8,43,400,40,38,9,10,57])
    dis = wb.create_sheet('VTHC_Doses disaggregate')
    dis.append(['Year','period','Organization','Project Name','District (EHO)','Township_EHO','Twp_MIMU','Clinic Name','ALOD-U1','ALOD-U5','ALOD->5','BCG_U1','BCG_U5','BCG_>5','OPV1_U1','OPV1_U5','OPV1_>5','OPV2_U1','OPV2_U5','OPV2_>5','OPV3_U1','OPV3_U5','OPV3_>5','Penta1_U1','Penta1_U5','Penta1_>5','Penta2_U1','Penta2_U5','Penta2_>5','Penta3_U1','Penta3_U5','Penta3_>5','MMR1_U1','MMR1_U5','MMR1_>5','MMR2_U1','MMR2_U5','MMR2_>5','JE_U1','JE_U5','JE_>5','IPV_U1','IPV_U5','IPV_>5','CD_U1','CD_U5','CD_>5','Td1','Td2','Td At least one dose'])
    dis.append(['2024','Q1_2024','KDHW','REACH-KK','Bleet Dawei','Htee Mo Pwah (special area)','Dawei','Ah Mala Hta',1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46])
    wb.save(input_path)
    build_semester_report(input_path, output_path, source_sheet='Indicator', target_sheet='Semester_Report')
    out_wb = load_workbook(output_path)
    print(out_wb.sheetnames)
