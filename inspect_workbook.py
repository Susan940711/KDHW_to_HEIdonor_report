from openpyxl import load_workbook
from pathlib import Path

for path_str in ['temp_quarterly_compile.xlsx', 'temp_quarterly_compile_semester_report.xlsx']:
    path = Path(path_str)
    wb = load_workbook(path, data_only=False)
    print(path.name)
    print('sheets', wb.sheetnames)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print('---', sheet, '---')
        for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
            print(row)
        print()
